from openpyxl import Workbook
from openpyxl.styles import PatternFill
from ..schemas import ProfileDownload
async def create_document(file_name: str, data: ProfileDownload):
    wb = Workbook()
    ws = wb.active
    
    ws['A1'] = "Name"
    ws['A2'] = "Gender"
    ws['A3'] = "Race"
    
    ws['A6'] = "Total"
    ws['A7'] = "Front"
    ws['A8'] = "Side"
    
    ws['B5'] = "Score"
    ws['C5'] = "Max Score"
    ws['D5'] = "Percentage"
    
    ws['G5'] = "Front Profile"
    ws['G6'] = "No"
    ws['H6'] = "Feature"
    ws['I6'] = "Value"
    ws['J6'] = "Score"
    ws['K6'] = "Ideal Range"
    ws['L6'] = "Note"
    ws['M6'] = "Advice"
    
    ws['P5'] = "Side Profile"
    ws['P6'] = "No"
    ws['Q6'] = "Feature"
    ws['R6'] = "Value"
    ws['S6'] = "Score"
    ws['T6'] = "Ideal Range"
    ws['U6'] = "Note"
    ws['V6'] = "Advice"
    
    ws['C7'] = 304.5
    ws['C8'] = 195.5
    ws['C6'] = '=C7+C8'
    ws['B6'] = '=B7+B8'
    ws['B7'] = '=SUM(J7:J29)'
    ws['B8'] = '=SUM(S7:S28)'
    ws['D6'] = '=B6/C6*100'
    ws['D7'] = '=B7/C7*100'
    ws['D8'] = '=B8/C8*100'
    
    ws['B1'] = data.name
    ws['B2'] = data.gender
    ws['B3'] = data.race
    
    for index, feature in enumerate(data.features):
        if index < 23:
            no = index
            start_cell = 'G'
        else:
            no = index - 23
            start_cell = 'P'
        ws[chr(ord(start_cell)) + str(7 + no)] = no + 1
        ws[chr(ord(start_cell) + 1) + str(7 + no)] = feature["name"]
        ws[chr(ord(start_cell) + 1) + str(7 + no)].hyperlink = feature["image"]
        ws[chr(ord(start_cell) + 2) + str(7 + no)] = feature["value"]
        start = int('FF0000', 16)
        end = int('00FF00', 16)
        left = feature["score"] - feature["minscore"]
        right = feature["maxscore"] - feature["score"]
        val = (int(start) * int(right) + int(end) * int(left)) // (int(right) + int(left))
        col = format(val, '06X')
        fill = PatternFill(start_color=col, end_color=col, fill_type='solid')
        ws[chr(ord(start_cell) + 3) + str(7 + no)].fill = fill
        ws[chr(ord(start_cell) + 3) + str(7 + no)] = feature["score"]
        ws[chr(ord(start_cell) + 4) + str(7 + no)] = feature["ideal"]
        ws[chr(ord(start_cell) + 5) + str(7 + no)] = feature["meaning"]
        ws[chr(ord(start_cell) + 6) + str(7 + no)] = feature["advice"]
    
    wb.save(filename=file_name)
    